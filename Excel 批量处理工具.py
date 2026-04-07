import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 全局变量：存储选中的文件路径
selected_files = []
# 全局变量：拆分列名
split_column_name = ""


# ====================== 核心功能函数 ======================
def select_files():
    """选择多个Excel文件"""
    files = filedialog.askopenfilenames(
        title="选择Excel文件",
        filetypes=[("Excel文件", "*.xlsx;*.xls"), ("所有文件", "*.*")]
    )
    if files:
        global selected_files
        selected_files = list(files)
        file_list.delete(0, tk.END)
        for file in files:
            file_list.insert(tk.END, file)
        messagebox.showinfo("成功", f"已选中 {len(files)} 个文件")


def clear_files():
    """清空选中的文件"""
    global selected_files
    selected_files = []
    file_list.delete(0, tk.END)
    messagebox.showinfo("成功", "已清空文件列表")


def merge_excel():
    """批量合并多个Excel文件（相同表头）"""
    if not selected_files:
        messagebox.showwarning("提示", "请先选择要合并的Excel文件！")
        return

    try:
        # 读取所有文件并合并
        df_list = []
        for file in selected_files:
            df = pd.read_excel(file)
            # 可选：添加来源文件名列，便于追溯
            df["source_file"] = os.path.basename(file)
            df_list.append(df)

        # 合并数据（按行堆叠，重置索引）
        merged_df = pd.concat(df_list, ignore_index=True)

        # 保存合并后的文件
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile="合并后的总表.xlsx"
        )
        if not save_path:
            return

        merged_df.to_excel(save_path, index=False)
        messagebox.showinfo(
            "成功",
            f"✅ 合并完成！\n\n共合并 {len(selected_files)} 个文件\n总数据 {merged_df.shape[0]} 行 {merged_df.shape[1]} 列\n保存路径：{save_path}"
        )
    except Exception as e:
        messagebox.showerror("失败", f"合并失败：{str(e)}")


def split_excel():
    """批量拆分Excel文件（按指定列拆分）"""
    if len(selected_files) != 1:
        messagebox.showwarning("提示", "请选择**一个**要拆分的Excel文件！")
        return

    # 选择拆分列
    split_window = tk.Toplevel(root)
    split_window.title("选择拆分列")
    split_window.geometry("300x200")
    split_window.config(bg="#f8f9fa")
    split_window.grab_set()

    tk.Label(
        split_window, text="选择拆分列名：",
        font=("微软雅黑", 12), bg="#f8f9fa"
    ).pack(pady=20)

    # 读取文件获取列名
    try:
        df = pd.read_excel(selected_files[0])
        columns = df.columns.tolist()
    except Exception as e:
        messagebox.showerror("错误", f"读取文件失败：{str(e)}")
        return

    # 列名选择下拉框
    col_var = tk.StringVar(value=columns[0] if columns else "")
    col_menu = ttk.OptionMenu(split_window, col_var, *columns)
    col_menu.pack(pady=10)

    def confirm_split():
        global split_column_name
        split_column_name = col_var.get()
        split_window.destroy()
        _perform_split(selected_files[0], split_column_name)

    tk.Button(
        split_window, text="确定拆分", command=confirm_split,
        font=("微软雅黑", 11, "bold"), bg="#1677ff", fg="white",
        relief="flat", width=15
    ).pack(pady=20)


def _perform_split(file_path, split_col):
    """执行拆分逻辑"""
    try:
        df = pd.read_excel(file_path)
        if split_col not in df.columns:
            messagebox.showerror("错误", f"列名 {split_col} 不存在！")
            return

        # 获取拆分列的唯一值
        unique_values = df[split_col].unique()
        total = len(unique_values)

        # 创建输出文件夹
        output_dir = filedialog.askdirectory(title="选择拆分结果保存文件夹")
        if not output_dir:
            return

        # 遍历分组并保存
        for i, value in enumerate(unique_values):
            sub_df = df[df[split_col] == value]
            # 文件名：拆分列值.xlsx
            save_name = f"{value}.xlsx"
            save_path = os.path.join(output_dir, save_name)
            sub_df.to_excel(save_path, index=False)

            # 更新进度
            progress = (i + 1) / total * 100
            status_label.config(text=f"拆分进度：{i + 1}/{total} ({progress:.1f}%)")
            root.update_idletasks()

        status_label.config(text="拆分完成！")
        messagebox.showinfo("成功", f"✅ 拆分完成！共生成 {total} 个文件，保存于：{output_dir}")
    except Exception as e:
        messagebox.showerror("失败", f"拆分失败：{str(e)}")


def batch_modify_header():
    """批量修改表头"""
    if not selected_files:
        messagebox.showwarning("提示", "请先选择要修改的Excel文件！")
        return

    # 表头修改窗口
    header_window = tk.Toplevel(root)
    header_window.title("批量修改表头")
    header_window.geometry("500x300")
    header_window.config(bg="#f8f9fa")
    header_window.grab_set()

    # 原表头输入
    tk.Label(
        header_window, text="原表头（多个用英文逗号分隔）：",
        font=("微软雅黑", 11), bg="#f8f9fa"
    ).pack(pady=10)
    old_header_entry = tk.Entry(header_window, width=50, font=("微软雅黑", 10))
    old_header_entry.pack(pady=5)

    # 新表头输入
    tk.Label(
        header_window, text="新表头（多个用英文逗号分隔）：",
        font=("微软雅黑", 11), bg="#f8f9fa"
    ).pack(pady=10)
    new_header_entry = tk.Entry(header_window, width=50, font=("微软雅黑", 10))
    new_header_entry.pack(pady=5)

    def confirm_modify():
        old_headers = [h.strip() for h in old_header_entry.get().split(",") if h.strip()]
        new_headers = [h.strip() for h in new_header_entry.get().split(",") if h.strip()]

        if len(old_headers) != len(new_headers):
            messagebox.showerror("错误", "原表头和新表头数量必须一致！")
            return

        try:
            for file in selected_files:
                df = pd.read_excel(file)
                # 重命名表头
                rename_dict = dict(zip(old_headers, new_headers))
                df = df.rename(columns=rename_dict)
                # 保存修改
                df.to_excel(file, index=False)

            messagebox.showinfo("成功", f"✅ 已修改 {len(selected_files)} 个文件的表头！")
            header_window.destroy()
        except Exception as e:
            messagebox.showerror("失败", f"修改失败：{str(e)}")

    tk.Button(
        header_window, text="确定修改", command=confirm_modify,
        font=("微软雅黑", 11, "bold"), bg="#1677ff", fg="white",
        relief="flat", width=15
    ).pack(pady=20)


def batch_format_unify():
    """批量统一格式（字体、对齐、填充）"""
    if not selected_files:
        messagebox.showwarning("提示", "请先选择要格式化的Excel文件！")
        return

    try:
        # 统一字体和对齐样式
        font = Font(name="微软雅黑", size=11, bold=False)
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for file in selected_files:
            wb = load_workbook(file)
            for ws in wb.worksheets:
                # 遍历所有单元格应用样式
                for row in ws.iter_rows():
                    for cell in row:
                        cell.font = font
                        cell.alignment = alignment
                        cell.fill = fill
                        cell.border = border
            wb.save(file)
            wb.close()

        messagebox.showinfo("成功", f"✅ 已统一 {len(selected_files)} 个文件的格式！")
    except Exception as e:
        messagebox.showerror("失败", f"格式化失败：{str(e)}")


# ====================== GUI 界面设计 ======================
root = tk.Tk()
root.title("Excel批量处理工具（合并/拆分/改表头/统格式）")
root.geometry("800x600")
root.resizable(False, False)
root.config(bg="#f8f9fa")

# 标题
tk.Label(
    root, text="Excel 批量处理工具",
    font=("微软雅黑", 20, "bold"),
    bg="#f8f9fa", fg="#2c3e50"
).pack(pady=20)

# 文件选择区域
frame_select = tk.Frame(root, bg="#f8f9fa")
frame_select.pack(pady=10, fill=tk.X, padx=20)

tk.Label(
    frame_select, text="已选文件：",
    font=("微软雅黑", 12), bg="#f8f9fa"
).pack(side=tk.LEFT, padx=5)

file_list = tk.Listbox(frame_select, width=80, height=5, font=("Consolas", 10))
file_list.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

frame_btn = tk.Frame(frame_select, bg="#f8f9fa")
frame_btn.pack(side=tk.LEFT, padx=10)

tk.Button(
    frame_btn, text="选择文件", command=select_files,
    font=("微软雅黑", 11, "bold"), bg="#1677ff", fg="white",
    relief="flat", width=12
).pack(pady=5)

tk.Button(
    frame_btn, text="清空列表", command=clear_files,
    font=("微软雅黑", 11, "bold"), bg="#dc3545", fg="white",
    relief="flat", width=12
).pack(pady=5)

# 功能按钮区域
frame_func = tk.Frame(root, bg="#f8f9fa")
frame_func.pack(pady=20, fill=tk.X, padx=20)

# 第一行按钮
tk.Button(
    frame_func, text="1. 批量合并Excel", command=merge_excel,
    font=("微软雅黑", 12, "bold"), bg="#28a745", fg="white",
    relief="flat", width=20, height=2
).grid(row=0, column=0, padx=10, pady=5)

tk.Button(
    frame_func, text="2. 批量拆分Excel", command=split_excel,
    font=("微软雅黑", 12, "bold"), bg="#ff9800", fg="white",
    relief="flat", width=20, height=2
).grid(row=0, column=1, padx=10, pady=5)

# 第二行按钮
tk.Button(
    frame_func, text="3. 批量修改表头", command=batch_modify_header,
    font=("微软雅黑", 12, "bold"), bg="#9c27b0", fg="white",
    relief="flat", width=20, height=2
).grid(row=1, column=0, padx=10, pady=5)

tk.Button(
    frame_func, text="4. 批量统格式", command=batch_format_unify,
    font=("微软雅黑", 12, "bold"), bg="#00acc1", fg="white",
    relief="flat", width=20, height=2
).grid(row=1, column=1, padx=10, pady=5)

# 状态显示区域
frame_status = tk.Frame(root, bg="#f8f9fa")
frame_status.pack(pady=10, fill=tk.X, padx=20)

tk.Label(
    frame_status, text="运行状态：",
    font=("微软雅黑", 12), bg="#f8f9fa", fg="#666"
).pack(side=tk.LEFT, padx=5)

status_label = tk.Label(
    frame_status, text="准备就绪",
    font=("微软雅黑", 12), bg="#f8f9fa", fg="#28a745"
)
status_label.pack(side=tk.LEFT, padx=5)

# 底部说明
tk.Label(
    root, text="💡 说明：支持.xlsx/.xls格式；合并为相同表头拼接，拆分为按指定列生成独立文件；修改表头/统格式可批量批量生效",
    font=("微软雅黑", 9), bg="#f8f9fa", fg="#666"
).pack(pady=15)

# 启动主循环
root.mainloop()